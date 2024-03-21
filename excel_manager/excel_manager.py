# excel_manager/excel_manager.py

"""This module allows the user to make excel operations.

Examples:
    >>> from excel_manager.excel_manager import ExcelManager
    >>> excel_manager = ExcelManager(log_file='abc.log')

    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> dataframe = excel_manager.get_dataframe(workbook, sheet)

    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> excel_manager.delete_sheet(workbook, sheet)

    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> excel_manager.create_sheet(workbook, sheet)

    >>> import pandas as pd
    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> dataframe = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
    >>> excel_manager.overwrite_sheet(workbook, sheet, dataframe)

    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> excel_manager.reposition_sheet(workbook, sheet)

    >>> import pandas as pd
    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> dataframe = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
    >>> excel_manager.append_dataframe(workbook, sheet, dataframe)

    >>> # to set the protection, 'abc' is password.
    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> excel_manager.modify_sheet_protection(workbook, sheet, True, 'abc')

    >>> # to remove the protection.
    >>> workbook = "test.xlsx"
    >>> sheet = "Sheet1"
    >>> excel_manager.modify_sheet_protection(workbook, sheet, False)

The module contains the following functions:

- `__init__(log_file)` - creates the instance of the class.
- `get_dataframe(workbook, sheet)` - returns the dataframe from excel workbook.
- `delete_sheet(workbook, sheet)` - deletes the sheet from excel workbook.
- `create_sheet(workbook, sheet)` - creates a new sheet in excel workbook.
- `overwrite_sheet(workbook, sheet, dataframe)` - overwrites the content of the sheet.
- `reposition_sheet(workbook, sheet)` - reposition the sheet in excel workbook.
- `append_dataframe(workbook, sheet, dataframe)` - append the dataframe to the content in excel sheet.
- `modify_sheet_protection(workbook, sheet, True, 'abc')` - adds or removes the protection from sheet in excel workbook.

"""

from pathlib import Path
import pandas as pd
import openpyxl
from date_manager import DateManager
from log_manager import LogManager
import inspect


class ExcelManager:
    def __init__(self, log_file: str = './Custom-Python_Tools.log') -> None:
        """
        Args:
            log_file: The path to the log file.
        """
        self.log = LogManager(log_name='ExcelManager', log_file=log_file)
        self.obj_date = DateManager(log_file=log_file)
        self.log.info("ExcelManager Initialized.")

    def get_dataframe(self, workbook: str, sheet: str) -> pd.DataFrame:
        """Retrieve a pandas dataframe from an Excel workbook.

        Args:
            workbook: The path to the Excel workbook.
            sheet: The name of the sheet containing the data.

        Returns:
            pd.DataFrame: The contents of the specified sheet as a pandas dataframe.

        Raises:
            FileNotFoundError: If the specified workbook cannot be found.
            Exception: If an unexpected error occurs.
        """
        try:
            self.log.info(f"{inspect.stack()[0][3]}:\nworkbook={workbook}, sheet={sheet}.")
            workbook_path = Path(workbook)
            if not workbook_path.is_file():
                raise FileNotFoundError
            dataframe = pd.read_excel(workbook_path, sheet_name=sheet)
            dataframe = self.obj_date.timestamp_to_date(dataframe)
            self.log.info("Dataframe created and timestamp columns modified.")
            return dataframe
        except FileNotFoundError:
            self.log.error(f"File '{workbook}' not found")
            self.log.warning("Initializing empty dataframe.")
            return pd.DataFrame()
        except Exception as e:
            self.log.error(f"Undefined Error: {e}.")
            self.log.warning("Initializing empty dataframe.")
            return pd.DataFrame()

    def delete_sheet(self, workbook: str, sheet: str) -> None:
        """Delete a sheet from an Excel workbook.

        Args:
            workbook: The path to the Excel workbook.
            sheet: The name of the sheet to be deleted.

        Returns:
            None: Returns nothing.

        Raises:
            FileNotFoundError: If the specified workbook cannot be found.
            ValueError: If the specified sheet does not exist in the workbook.
            Exception: If an unexpected error occurs.
        """
        self.log.info(f"{inspect.stack()[0][3]}:\nworkbook={workbook}, sheet={sheet}.")
        try:
            if not Path(workbook).is_file():
                raise FileNotFoundError(f"File '{workbook}' not found.")
            wb = openpyxl.load_workbook(Path(workbook))
            sheet_names_list = wb.sheetnames
            self.log.info(f"{workbook} loaded and sheet names: {sheet_names_list}.")
            if sheet not in sheet_names_list:
                raise ValueError(f"{sheet} not present in {workbook}.")
            sheetno = sheet_names_list.index(sheet)
            del wb[sheet_names_list[sheetno]]
            wb.save(workbook)
            self.log.info(f"{sheet} deleted from {workbook}.")
        except (FileNotFoundError, ValueError) as e:
            self.log.error(e)
        except Exception as e:
            self.log.error(f"Undefined Error: {e}.")

    def create_sheet(self, workbook: str, sheet: str) -> None:
        """Create a new sheet in an Excel workbook.

        Args:
            workbook: The path to the Excel workbook.
            sheet: The name of the sheet to be created.

        Raises:
            FileNotFoundError: If the specified workbook cannot be found.
            PermissionError: If the user does not have permission to write to the specified workbook.
            Exception: If an unexpected error occurs.

        Returns:
            None: Returns nothing.
        """
        self.log.info(f"{inspect.stack()[0][3]}:\nworkbook={workbook}, sheet={sheet}.")
        try:
            if not Path(workbook).is_file():
                raise FileNotFoundError
            wb = openpyxl.load_workbook(workbook)
            wb.create_sheet(title=sheet)
            wb.save(workbook)
            self.log.info(f"{sheet} created in {workbook}.")
        except FileNotFoundError:
            self.log.error(f"File {workbook} not found.")
        except PermissionError:
            self.log.error(f"Permission error while creating sheet in {workbook}.")
        except Exception as e:
            self.log.error(f"Undefined Error: {e}.")

    def overwrite_sheet(self, workbook: str, sheet: str, dataframe: pd.DataFrame) -> None:
        """Overwrite the contents of an Excel sheet with a new dataframe.

        Args:
            workbook: The path to the Excel workbook.
            sheet: The name of the sheet to be overwritten.
            dataframe: The new contents of the sheet as a pandas dataframe.

        Returns:
            None: Returns nothing.

        Raises:
            FileNotFoundError: If the specified workbook cannot be found.
            PermissionError: If the user does not have permission to write to the specified workbook.
            ValueError: If the specified sheet does not exist in the workbook.
            Exception: If an unexpected error occurs.
        """
        self.log.info(f"{inspect.stack()[0][3]}:\nworkbook={workbook}, sheet={sheet}, dataframe_size={dataframe.shape}.")
        try:
            header = dataframe.columns
            self.delete_sheet(workbook=workbook, sheet=sheet)
            self.create_sheet(workbook=workbook, sheet=sheet)

            with pd.ExcelWriter(
                Path(workbook), mode="a", engine="openpyxl", if_sheet_exists="overlay"
            ) as writer:
                dataframe.to_excel(
                    writer, sheet_name=sheet, header=header, startrow=0, index=False
                )
            self.log.info(f"{sheet} overwritten in {workbook}.")
        except (FileNotFoundError, ValueError) as e:
            self.log.error(e)
        except PermissionError:
            self.log.error(f"Permission error while updating {workbook}.")
        except Exception as e:
            self.log.error(f"Undefined Error: {e}.")

    def reposition_sheet(self, workbook: str, sheet: str) -> None:
        """Moves a sheet at the start in the workbook and saves the changes.

        Args:
            workbook: The path to the Excel workbook.
            sheet: The name of the sheet to be moved.

        Raises:
            FileNotFoundError: If the specified workbook cannot be found.
            PermissionError: If the user does not have permission to write to the specified workbook.
            ValueError: If the specified sheet does not exist in the workbook.
            Exception: If an unexpected error occurs.

        Returns:
            None: Returns nothing.
        """
        self.log.info(f"{inspect.stack()[0][3]}:\nworkbook={workbook}, sheet={sheet}.")
        try:
            wb = openpyxl.load_workbook(workbook)
            if sheet not in wb.sheetnames:
                raise ValueError
            wb.move_sheet(sheet, -(len(wb.sheetnames) - 1))
            wb.save(workbook)
            wb.close()
            self.log.info(f"{sheet} moved to the start of {workbook}.")
        except FileNotFoundError:
            self.log.error(f"File '{workbook}' not found.")
        except PermissionError:
            self.log.error(f"Permission error while accessing '{workbook}'.")
        except ValueError:
            self.log.error(f"Sheet '{sheet}' not found in workbook '{workbook}'.")
        except Exception as e:
            self.log.error(f"Undefined Error: {e}.")

    def append_dataframe(self, workbook: str, sheet: str, dataframe: pd.DataFrame, password: str = None) -> None:
        """Appends a pandas dataframe to an Excel sheet.

        Args:
            workbook: The path to the Excel workbook.
            sheet: The name of the sheet to which the dataframe will be appended.
            dataframe: The pandas dataframe to be appended to the sheet.
            password: The password for the Excel workbook, if it is protected.

        Raises:
            FileNotFoundError: If the specified workbook cannot be found.
            PermissionError: If the user does not have permission to write to the specified workbook.
            Exception: If an unexpected error occurs.

        Returns:
            None: Returns nothing.
        """
        self.log.info(f"{inspect.stack()[0][3]}:\nworkbook={workbook}, sheet={sheet}, dataframe={dataframe.shape}.")
        try:
            if password:
                self.modify_sheet_protection(filepath=workbook, sheetname=sheet, enable_protection=False, password=password)
            with pd.ExcelWriter(
                Path(workbook), mode="a", engine="openpyxl", if_sheet_exists="overlay"
            ) as writer:
                dataframe.to_excel(
                    writer,
                    sheet_name=sheet,
                    header=None,
                    startrow=writer.sheets[sheet].max_row if sheet in writer.sheets else 0,
                    index=False,
                )
            self.log.info(f"Dataframe appended to {sheet} in {workbook}.")
            if password:
                self.modify_sheet_protection(filepath=workbook, sheetname=sheet, enable_protection=True, password=password)
        except FileNotFoundError:
            self.log.error(f"File '{workbook}' not found.")
        except PermissionError:
            self.log.error(f"Permission error while accessing '{workbook}'.")
        except Exception as e:
            self.log.error(f"Undefined Error: {e}.")

    def modify_sheet_protection(self, filepath: str, sheetname: str, enable_protection: bool, password: str = None) -> None:
        """Modifies the protection of an Excel sheet.

        Args:
            filepath: The path to the Excel workbook.
            sheetname: The name of the sheet to be protected.
            enable_protection: A boolean value indicating whether to enable or disable protection.
            password: The password for the Excel workbook, if it is protected.

        Raises:
            FileNotFoundError: If the specified workbook cannot be found.
            PermissionError: If the user does not have permission to write to the specified workbook.
            
        Returns:
            None: Returns nothing.
        """
        self.log.info(f"{inspect.stack()[0][3]}:\nworkbook={filepath}, sheet={sheetname}, enable_protection={enable_protection}.")
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheetname]
            ws.protection.sheet = enable_protection
            protect = 'enabled' if enable_protection else 'disabled'
            if enable_protection and password:
                ws.protection.password = password
            wb.save(filepath)
            wb.close()
            self.log.info(f"Protection {protect} for {sheetname} in {filepath}.")
        except FileNotFoundError:
            self.log.error(f"File '{filepath}' not found.")
        except PermissionError:
            self.log.error(f"Permission error while accessing '{filepath}'.")
        except Exception as e:
            self.log.error(f"Undefined Error: {e}.")
