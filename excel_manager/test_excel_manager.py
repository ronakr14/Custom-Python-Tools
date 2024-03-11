import unittest
import openpyxl
import pandas as pd
from pathlib import Path
from excel_manager import ExcelManager


class TestExcelManager(unittest.TestCase):

    def setUp(self):
        self.obj = ExcelManager()
        workbook = openpyxl.Workbook()
        workbook.save("test.xlsx")
        sheet = workbook.active
        sheet.title = "Sheet1"
        workbook.create_sheet(title="Sheet2")
        workbook.save("test.xlsx")

    def test_get_dataframe(self):
        workbook = "test.xlsx"
        sheet = "Sheet1"
        dataframe = self.obj.get_dataframe(workbook, sheet)
        self.assertIsInstance(dataframe, pd.DataFrame)

    def test_delete_sheet(self):
        workbook = "test.xlsx"
        sheet = "Sheet1"
        self.obj.delete_sheet(workbook, sheet)
        self.assertTrue(Path(workbook).is_file())

    def test_create_sheet(self):
        workbook = "test.xlsx"
        sheet = "Sheet1"
        self.obj.create_sheet(workbook, sheet)
        self.assertTrue(Path(workbook).is_file())

    def test_overwrite_sheet(self):
        workbook = "test.xlsx"
        sheet = "Sheet1"
        dataframe = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
        self.obj.overwrite_sheet(workbook, sheet, dataframe)
        dataframe_test = self.obj.get_dataframe(workbook, sheet)
        self.assertTrue(dataframe.equals(dataframe_test))

    def test_reposition_sheet(self):
        workbook = "test.xlsx"
        sheet = "Sheet1"
        self.obj.reposition_sheet(workbook, sheet)
        wb = openpyxl.load_workbook(workbook)
        self.assertEqual(sheet, wb.sheetnames[0])

    def test_append_dataframe(self):
        workbook = "test.xlsx"
        sheet = "Sheet1"
        dataframe = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
        self.obj.append_dataframe(workbook, sheet, dataframe)
        dataframe_test = self.obj.get_dataframe(workbook, sheet)
        dataframe_test = dataframe_test.rename(columns={'Unnamed: 0': 'a', 'Unnamed: 1': 'b'})
        self.assertTrue(dataframe.equals(dataframe_test))

    def test_modify_sheet_protection(self):
        workbook = "test.xlsx"
        sheet = "Sheet1"
        self.obj.modify_sheet_protection(workbook, sheet, True, 'abc')
        wb = openpyxl.load_workbook(workbook)
        ws = wb[sheet]
        self.assertTrue(ws.protection.sheet)
        self.obj.modify_sheet_protection(workbook, sheet, False)
        self.assertTrue(ws.protection.sheet)


if __name__ == '__main__':
    unittest.main()
