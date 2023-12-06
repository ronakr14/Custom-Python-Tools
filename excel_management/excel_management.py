#!/usr/bin/env python
# coding: utf-8

from pathlib import Path
import pandas as pd
import openpyxl
from date_management import Date_Management


class Excel_Management:
    def __init__(self):
        pass

    def get_dataframe(Workbook: str, Sheet: str):
        try:
            dataframe = pd.read_excel(Path(Workbook), sheet_name=Sheet)
            dataframe = Date_Management.timestamp_to_date(dataframe)
        except Exception:
            dataframe = pd.DataFrame()
        return dataframe

    def delete_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        sheet_names_list = wb.sheetnames
        sheetno = sheet_names_list.index(Sheet)
        del wb[sheet_names_list[sheetno]]
        wb.save(Workbook)

    def create_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.create_sheet(title=Sheet)
        wb.save(Workbook)

    def overwrite_sheet(Workbook: str, Sheet: str, dataframe):
        header = dataframe.columns
        Excel_Management.delete_excel_sheet(Workbook=Workbook, Sheet=Sheet)
        Excel_Management.create_sheet(Workbook=Workbook, Sheet=Sheet)
        with pd.ExcelWriter(
            Path(Workbook), mode="a", engine="openpyxl",
            if_sheet_exists="overlay"
        ) as writer:
            dataframe.to_excel(
                writer, sheet_name=Sheet, header=header, startrow=0,
                index=False
            )

    def reposition_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.move_sheet(Sheet, -(len(wb.sheetnames) - 1))
        wb.save(Workbook)
        wb.close()

    def append_dataframe_w_password(Workbook: str, Sheet: str,
                                    Password: str, dataframe):
        Excel_Management.remove_password_sheet(
            filepath=Workbook, sheetname=Sheet)
        with pd.ExcelWriter(
            Path(Workbook), mode="a", engine="openpyxl",
            if_sheet_exists="overlay"
        ) as writer:
            dataframe.to_excel(
                writer,
                sheet_name=Sheet,
                header=None,
                startrow=writer.sheets[Sheet].max_row,
                index=False,
            )

        Excel_Management.add_sheet_password(
            filepath=Workbook, sheetname=Sheet, password=Password
        )

    def append_dataframe_wo_password(Workbook: str, Sheet: str, dataframe):
        with pd.ExcelWriter(
            Path(Workbook), mode="a", engine="openpyxl",
            if_sheet_exists="overlay"
        ) as writer:
            dataframe.to_excel(
                writer,
                sheet_name=Sheet,
                header=None,
                startrow=writer.sheets[Sheet].max_row,
                index=False,
            )

    def remove_password_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.active = wb[Sheet]
        ws = wb.active
        ws.protection
        ws.protection.sheet = False
        wb.save(Workbook)
        wb.close()

    def add_sheet_password(Workbook: str, Sheet: str, Password: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.active = wb[Sheet]
        ws = wb.active
        ws.protection
        ws.protection.sheet = True
        ws.protection.password = Password
        wb.save(Workbook)
        wb.close()
